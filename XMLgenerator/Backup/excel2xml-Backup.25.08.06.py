#Argonne Charging Curve Spreadheet to XML Converter 
#By Aaron Sobel, EPA
#
# This script converts Argonne's Charing curve spreadsheets into an XML format that is compatible with the EV Road Trip Metric tool.
#The following python libraries are required to be installed:
# pip install pandas numpy openpyxl
# Place all the Argonne spreadshets in the inputFiles folder before running this script.
# The output XML files will be saved in the outputFiles folder.

import os
import sys
import pandas as pd
import numpy as np
from xml.etree.ElementTree import Element, SubElement, ElementTree

INPUT_DIR = os.path.join(os.path.dirname(__file__), 'inputFiles')
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), 'outputFiles')


def supports_color():
    # Returns True if the running terminal supports color
    if sys.platform == 'win32':
        return 'ANSICON' in os.environ or 'WT_SESSION' in os.environ or os.environ.get('TERM_PROGRAM') == 'vscode'
    if hasattr(sys.stdout, 'isatty') and sys.stdout.isatty():
        return True
    return False


def list_excel_files():
    import re
    def natural_key(s):
        # Split string into list of strings and integers for natural sorting
        return [int(text) if text.isdigit() else text.lower() for text in re.split(r'(\d+)', s)]
    files = [f for f in os.listdir(INPUT_DIR) if f.lower().endswith(('.xlsx', '.xls'))]
    files.sort(key=natural_key)
    if not files:
        print('No Excel files found in inputFiles folder.')
        sys.exit(1)
    print('Available Excel files:')
    for idx, fname in enumerate(files, 1):
        print(f'{idx}: {fname}')
    return files


def prompt_file_selection(files):
    while True:
        default_choice = 1
        resp = input(f'Select a file to convert (number) [{default_choice}]: ').strip()
        try:
            choice = int(resp) if resp else default_choice
            if 1 <= choice <= len(files):
                return files[choice - 1]
        except Exception:
            pass
        print('Invalid selection. Try again.')


def prompt_interpolation():
    default = 'y'
    while True:
        resp = input(f'Can I interpolate between values as needed? (y/n) [{default}]: ').strip().lower()
        if not resp:
            resp = default
        if resp in ('y', 'n'):
            return resp == 'y'
        print('Please enter y or n.')


def read_spreadsheet(filepath):
    try:
        # Always read from Sheet1, header in row 3, data starts at row 4
        df = pd.read_excel(filepath, sheet_name='Sheet1', header=2)
        # Only keep rows from row 4 onward (header=2 means pandas row 0 is Excel row 3)
        df = df.dropna(how='all')
        # Map columns by Excel column letters
        # CK: Battery Display SOC [%], CE: Power [W(DC)], H: Time [hh:mm:ss.0], CF: Energy Charged [Wh(DC)]
        col_soc = 'Battery Display SOC [%]'
        col_power = 'Power [W(DC)]'
        col_time = 'Time [hh:mm:ss.0]'
        col_energy = 'Energy Charged [Wh(DC)]'
        for col in [col_soc, col_power, col_time, col_energy]:
            if col not in df.columns:
                print(f'Missing required column: {col}')
                print(f'Columns found: {list(df.columns)}')
                sys.exit(1)
        # print(f"Read {len(df)} rows from Sheet1 with columns: {list(df.columns)}")
        return df[[col_soc, col_power, col_time, col_energy]].copy()
    except Exception as e:
        print(f'Error reading spreadsheet: {e}')
        sys.exit(1)


def process_data(df, interpolate, extrapolate):
    # Use fixed column names as per user specification
    col_soc = 'Battery Display SOC [%]'
    col_power = 'Power [W(DC)]'
    col_time = 'Time [hh:mm:ss.0]'
    col_energy = 'Energy Charged [Wh(DC)]'
    try:
        soc = df[col_soc].astype(float)
        power = df[col_power].astype(float) / 1000.0  # Convert W to kW
        def time_to_minutes(t):
            if pd.isnull(t):
                return np.nan
            parts = str(t).split(':')
            if len(parts) == 3:
                h, m, s = parts
                return float(h) * 60 + float(m) + float(s) / 60
            return np.nan
        time = df[col_time].apply(time_to_minutes)
        energy = df[col_energy].astype(float) / 1000.0  # Convert Wh to kWh
    except Exception as e:
        RED = '\033[91m'
        RESET = '\033[0m'
        print(f"{RED}Error: Could not convert one or more columns to numeric values.{RESET}")
        print(f"{RED}Please check that your spreadsheet contains valid data for the following columns: 'Battery Display SOC [%]', 'Power [W(DC)]', 'Time [hh:mm:ss.0]', and 'Energy Charged [Wh(DC)]'.{RESET}")
        print(f"{RED}Details: {e}{RESET}")
        input("Press Enter to exit...")
        sys.exit(1)
    # Scale SOC to 0-100 if needed
    if soc.max() <= 1.0:
        soc = soc * 100
    # Remove duplicate SOCs, keep last occurrence
    dedup = pd.DataFrame({'soc': soc, 'power': power, 'time': time, 'energy': energy})
    dedup = dedup.drop_duplicates(subset='soc', keep='last').sort_values('soc')
    soc = dedup['soc'].values
    power = dedup['power'].values
    time = dedup['time'].values
    energy = dedup['energy'].values

    # Extrapolation choice is now passed in
    soc_min, soc_max = soc[0], soc[-1]
    missing_low = int(soc_min) > 0
    missing_high = int(soc_max) < 100

    soc_grid = np.arange(0, 101)
    power_out = []
    time_out = []
    energy_out = []

    if not interpolate:
        # If extrapolate is True, extrapolate ends, but do not interpolate between points
        if extrapolate:
            # Extrapolate lower end
            if missing_low:
                first_soc = soc[0]
                n_missing = int(first_soc)
                usable_fraction = (100 - first_soc) / 100.0
                if usable_fraction > 0:
                    total_capacity = energy[-1] / usable_fraction
                else:
                    total_capacity = energy[-1]
                missing_energy = total_capacity - energy[-1]
                extra_soc = np.arange(0, int(first_soc))
                extra_energy = np.linspace(0, missing_energy, n_missing, endpoint=False)
                if len(energy) > 1 and len(time) > 1:
                    measured_dE = energy[1] - energy[0]
                    measured_dt = time[1] - time[0]
                    template_power = measured_dE / (measured_dt / 60) if measured_dt > 0 else power[0]
                else:
                    template_power = power[0]
                dE = np.diff(np.concatenate(([0], extra_energy)))
                extra_dt = np.where(template_power > 0, dE / template_power * 60, 0)
                extra_time = np.cumsum(extra_dt)
                with np.errstate(divide='ignore', invalid='ignore'):
                    extra_power = np.where(extra_dt > 0, dE / (extra_dt / 60), 0)
                    extra_power = np.nan_to_num(extra_power, nan=0.0, posinf=0.0, neginf=0.0)
                energy = energy + missing_energy
                total_extrapolated_time = extra_time[-1] if len(extra_time) > 0 else 0
                time = time + total_extrapolated_time
                soc = np.concatenate([extra_soc, soc])
                power = np.concatenate([extra_power, power])
                energy = np.concatenate([extra_energy, energy])
                time = np.concatenate([extra_time, time])
            # Extrapolate upper end
            if missing_high:
                last_soc = soc[-1]
                last_power = power[-1]
                last_energy = energy[-1]
                last_time = time[-1]
                n_missing = 100 - int(last_soc)
                if n_missing > 0:
                    if len(soc) > 1:
                        delta_e = energy[-1] - energy[-2]
                        delta_t = time[-1] - time[-2]
                    else:
                        delta_e = last_power / 60
                        delta_t = 1
                    extra_soc = np.arange(int(last_soc)+1, 101)
                    extra_power = np.full_like(extra_soc, last_power, dtype=float)
                    extra_energy = np.linspace(last_energy, last_energy + delta_e * n_missing, n_missing+1)[1:]
                    extra_time = np.linspace(last_time, last_time + delta_t * n_missing, n_missing+1)[1:]
                    soc = np.concatenate([soc, extra_soc])
                    power = np.concatenate([power, extra_power])
                    energy = np.concatenate([energy, extra_energy])
                    time = np.concatenate([time, extra_time])
            # Output only original data points for in-range SOCs, extrapolated values for out-of-range
            soc_map = {int(round(s)): i for i, s in enumerate(soc)}
            for s in soc_grid:
                if s < soc_min or s > soc_max:
                    # Use extrapolated value
                    idx = soc_map.get(s, None)
                    if idx is not None:
                        power_out.append(f'{power[idx]:.0f}')
                        time_out.append(f'{time[idx]:.8f}')
                        energy_out.append(f'{energy[idx]:.1f}')
                    else:
                        power_out.append('')
                        time_out.append('')
                        energy_out.append('')
                else:
                    # Only output original data points
                    idx = soc_map.get(s, None)
                    if idx is not None:
                        power_out.append(f'{power[idx]:.0f}')
                        time_out.append(f'{time[idx]:.8f}')
                        energy_out.append(f'{energy[idx]:.1f}')
                    else:
                        power_out.append('')
                        time_out.append('')
                        energy_out.append('')
            return soc_grid, np.array(power_out), np.array(time_out), np.array(energy_out)
        else:
            # Only output original data points, blanks elsewhere
            soc_map = {int(round(s)): i for i, s in enumerate(soc)}
            for s in soc_grid:
                idx = soc_map.get(s, None)
                if idx is not None:
                    power_out.append(f'{power[idx]:.0f}')
                    time_out.append(f'{time[idx]:.8f}')
                    energy_out.append(f'{energy[idx]:.1f}')
                else:
                    power_out.append('')
                    time_out.append('')
                    energy_out.append('')
            return soc_grid, np.array(power_out), np.array(time_out), np.array(energy_out)

    if extrapolate:
        # Extrapolate lower end: energy first, then time, then power (power = dE/dt for each step)
        if missing_low:
            first_soc = soc[0]
            n_missing = int(first_soc)
            usable_fraction = (100 - first_soc) / 100.0
            if usable_fraction > 0:
                total_capacity = energy[-1] / usable_fraction
            else:
                total_capacity = energy[-1]
            missing_energy = total_capacity - energy[-1]
            extra_soc = np.arange(0, int(first_soc))
            extra_energy = np.linspace(0, missing_energy, n_missing, endpoint=False)
            if len(energy) > 1 and len(time) > 1:
                measured_dE = energy[1] - energy[0]
                measured_dt = time[1] - time[0]
                template_power = measured_dE / (measured_dt / 60) if measured_dt > 0 else power[0]
            else:
                template_power = power[0]
            dE = np.diff(np.concatenate(([0], extra_energy)))
            extra_dt = np.where(template_power > 0, dE / template_power * 60, 0)
            extra_time = np.cumsum(extra_dt)
            with np.errstate(divide='ignore', invalid='ignore'):
                extra_power = np.where(extra_dt > 0, dE / (extra_dt / 60), 0)
                extra_power = np.nan_to_num(extra_power, nan=0.0, posinf=0.0, neginf=0.0)
            energy = energy + missing_energy
            total_extrapolated_time = extra_time[-1] if len(extra_time) > 0 else 0
            time = time + total_extrapolated_time
            soc = np.concatenate([extra_soc, soc])
            power = np.concatenate([extra_power, power])
            energy = np.concatenate([extra_energy, energy])
            time = np.concatenate([extra_time, time])
        # Extrapolate upper end
        if missing_high:
            last_soc = soc[-1]
            last_power = power[-1]
            last_energy = energy[-1]
            last_time = time[-1]
            n_missing = 100 - int(last_soc)
            if n_missing > 0:
                if len(soc) > 1:
                    delta_e = energy[-1] - energy[-2]
                    delta_t = time[-1] - time[-2]
                else:
                    delta_e = last_power / 60
                    delta_t = 1
                extra_soc = np.arange(int(last_soc)+1, 101)
                extra_power = np.full_like(extra_soc, last_power, dtype=float)
                extra_energy = np.linspace(last_energy, last_energy + delta_e * n_missing, n_missing+1)[1:]
                extra_time = np.linspace(last_time, last_time + delta_t * n_missing, n_missing+1)[1:]
                soc = np.concatenate([soc, extra_soc])
                power = np.concatenate([power, extra_power])
                energy = np.concatenate([energy, extra_energy])
                time = np.concatenate([time, extra_time])
        # Interpolate/extrapolate to integer SOC 0-100
        power_interp = np.interp(soc_grid, soc, power)
        time_interp = np.interp(soc_grid, soc, time)
        if energy is not None:
            energy_interp = np.interp(soc_grid, soc, energy)
        else:
            energy_interp = np.zeros_like(soc_grid, dtype=float)
            for i in range(1, len(soc_grid)):
                dt = time_interp[i] - time_interp[i-1]
                avg_power = (power_interp[i] + power_interp[i-1]) / 2
                energy_interp[i] = energy_interp[i-1] + avg_power * dt / 60
        return soc_grid, power_interp, time_interp, energy_interp

    # Interpolate only within bounds, blanks outside
    for s in soc_grid:
        if soc_min <= s <= soc_max:
            power_val = float(np.interp(s, soc, power))
            time_val = float(np.interp(s, soc, time))
            energy_val = float(np.interp(s, soc, energy)) if energy is not None else ''
            power_out.append(f'{power_val:.0f}')
            time_out.append(f'{time_val:.8f}')
            energy_out.append(f'{energy_val:.1f}' if energy_val != '' else '')
        else:
            power_out.append('')
            time_out.append('')
            energy_out.append('')
    return soc_grid, np.array(power_out), np.array(time_out), np.array(energy_out)

   


def build_xml(soc, power, time, energy, vehicle_name, vehicle_range, range_speed):
    vehicle = Element('vehicle')
    metadata = SubElement(vehicle, 'metadata')
    SubElement(metadata, 'name').text = vehicle_name
    SubElement(metadata, 'range_miles').text = str(vehicle_range)
    SubElement(metadata, 'range_speed_mph').text = str(range_speed)
    curve = SubElement(vehicle, 'charging_curve', unit_time='minutes', unit_power='kW', unit_energy='kWh')
    for s, p, t, e in zip(soc, power, time, energy):
        # If any value is blank, output blank string, else format as float
        power_str = '' if p == '' else f'{float(p):.0f}'
        time_str = '' if t == '' else f'{float(t):.2f}'
        energy_str = '' if e == '' else f'{float(e):.1f}'
        SubElement(curve, 'point', soc=str(int(s)), power=power_str, time=time_str, energy=energy_str)
    return vehicle


def prompt_output_format():
    print("Select output format:")
    print("1: XML (default)")
    print("2: CSV")
    default_choice = 1
    while True:
        resp = input(f"Choose output format (number) [{default_choice}]: ").strip()
        try:
            choice = int(resp) if resp else default_choice
            if choice in (1, 2):
                return 'xml' if choice == 1 else 'csv'
        except Exception:
            pass
        print('Invalid selection. Try again.')

def prompt_csv_format():
    print("Select CSV output format:")
    print("1: Data-cleaned (default, no units in values)")
    print("2: EVKX-format (As data appears on evkx.net)")
    default_choice = 1
    while True:
        resp = input(f"Choose CSV format (number) [{default_choice}]: ").strip()
        try:
            choice = int(resp) if resp else default_choice
            if choice in (1, 2):
                return 'cleaned' if choice == 1 else 'evkx'
        except Exception:
            pass
        print('Invalid selection. Try again.')


def main():
    files = list_excel_files()
    fname = prompt_file_selection(files)
    output_format = prompt_output_format()
    csv_format = None
    if output_format == 'csv':
        csv_format = prompt_csv_format()
    df = read_spreadsheet(os.path.join(INPUT_DIR, fname))
    import openpyxl
    wb = openpyxl.load_workbook(os.path.join(INPUT_DIR, fname), data_only=True)
    ws = wb.active
    detected_name = ws['B3'].value if ws['B3'].value else ''
    if detected_name:
        print(f"Detected vehicle name from spreadsheet: {detected_name}")
        vehicle_name = input(f"Vehicle name [{detected_name}]: ").strip()
        if not vehicle_name:
            vehicle_name = detected_name
    else:
        default_vehicle = ''
        vehicle_name = input(f'Vehicle name [{default_vehicle}]: ').strip()
        if not vehicle_name:
            vehicle_name = default_vehicle
    vehicle_range = input('Range (miles): ')
    interpolate = prompt_interpolation()
    range_speed = 70
    # Prompt for extrapolation BEFORE processing data
    # Determine if extrapolation is needed by inspecting the raw data
    col_soc = 'Battery Display SOC [%]'
    soc_raw = df[col_soc].astype(float)
    if soc_raw.max() <= 1.0:
        soc_raw = soc_raw * 100
    soc_raw = soc_raw.sort_values()
    soc_min, soc_max = soc_raw.iloc[0], soc_raw.iloc[-1]
    missing_low = int(soc_min) > 0
    missing_high = int(soc_max) < 100
    if missing_low or missing_high:
        YELLOW = '\033[93m'
        RESET = '\033[0m'
        msg = []
        if missing_low:
            msg.append(f"{int(soc_min)}% and below")
        if missing_high:
            msg.append(f"{int(soc_max)}% and above")
        print(f"{YELLOW}Data missing for SOC: {', '.join(msg)}.{RESET}")
        default = 'y'
        resp = input(f"Should I extrapolate outside of the provided data to give a fuller picture? (y/n) [{default}]: ").strip().lower()
        if not resp:
            resp = default
        extrapolate = (resp == 'y')
    else:
        extrapolate = False
    soc, power, time, energy = process_data(df, interpolate, extrapolate)
    import re
    safe_vehicle_name = vehicle_name.strip().replace('>', 'm').replace('<', 'l')
    safe_vehicle_name = re.sub(r'[^a-zA-Z0-9_\-]', '_', safe_vehicle_name)
    interp_suffix = '_interp' if interpolate else ''
    extrap_suffix = '_extrap' if extrapolate else ''
    filename_suffix = f"{interp_suffix}{extrap_suffix}" if interp_suffix or extrap_suffix else ''
    output_dir = os.path.join(os.path.dirname(__file__), 'outputFiles')
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    if output_format == 'xml':
        vehicle_xml = build_xml(soc, power, time, energy, safe_vehicle_name, vehicle_range, range_speed)
        output_xml = ElementTree(vehicle_xml)
        output_file = os.path.join(output_dir, f"{safe_vehicle_name}{filename_suffix}.xml")
        with open(output_file, 'wb') as f:
            output_xml.write(f)
        GREEN = '\033[92m'
        RESET = '\033[0m'
        print(f"{GREEN}Success: Generated XML file: {output_file}{RESET}")
        input("Press Enter to exit...")
        sys.exit(0)
    else:
        output_file = os.path.join(output_dir, f"{safe_vehicle_name}{filename_suffix}.csv")
        try:
            with open(output_file, 'w', encoding='utf-8', newline='') as f:
                # Write metadata rows
                f.write(f'name,{vehicle_name},,\n')
                f.write(f'range_miles,{vehicle_range},,\n')
                f.write(f'range_speed_mph,{range_speed},,\n')
                f.write('Battery State of Charge (%),Charging Speed (kW),"Time\n(Min)",Energy Charged (kWh)\n')
                # Write data rows
                for s, p, t, e in zip(soc, power, time, energy):
                    soc_pct = s / 100.0
                    if csv_format == 'evkx':
                        # EVKX-format: add units, time/1440
                        p_str = '' if p == '' else f'{float(p):.0f} kW'
                        t_str = '' if t == '' else f'{float(t)/1440:.8f}'
                        e_str = '' if e == '' else f'{float(e):.1f} kWh'
                    else:
                        # Data-cleaned: no units, time in minutes
                        p_str = '' if p == '' else f'{float(p):.0f}'
                        t_str = '' if t == '' else f'{float(t):.8f}'
                        e_str = '' if e == '' else f'{float(e):.1f}'
                    f.write(f'{soc_pct:.2f},{p_str},{t_str},{e_str}\n')
            GREEN = '\033[92m'
            RESET = '\033[0m'
            print(f"{GREEN}Success: Generated CSV file: {output_file}{RESET}")
            input("Press Enter to exit...")
            sys.exit(0)
        except PermissionError:
            RED = '\033[91m'
            RESET = '\033[0m'
            print(f"{RED}Error: Permission denied when writing CSV file. Please close the file if it is open in another program and try again.{RESET}")
            input("Press Enter to exit...")

if __name__ == '__main__':
    main()
