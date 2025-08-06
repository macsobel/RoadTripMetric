#Argonne Charging Curve Spreadheet to XML Converter 
#By Aaron Sobel, EPA
#
# This script converts Argonne's Charing curve spreadsheets into an XML format that is compatible with the EV Road Trip Metric tool.
#The following python libraries are required to be installed:
# pip install pandas numpy openpyxl
# Place all the Argonne spreadshets in the inputFiles folder before running this script.
# The output XML files will be saved in the outputFiles folder.

import os
import sys
import pandas as pd
import numpy as np
from xml.etree.ElementTree import Element, SubElement, ElementTree

INPUT_DIR = os.path.join(os.path.dirname(__file__), 'inputFiles')
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), 'outputFiles')


def list_excel_files():
    files = [f for f in os.listdir(INPUT_DIR) if f.lower().endswith(('.xlsx', '.xls'))]
    if not files:
        print('No Excel files found in inputFiles folder.')
        sys.exit(1)
    print('Available Excel files:')
    for idx, fname in enumerate(files, 1):
        print(f'{idx}: {fname}')
    return files


def prompt_file_selection(files):
    while True:
        default_choice = 1
        resp = input(f'Select a file to convert (number) [{default_choice}]: ').strip()
        try:
            choice = int(resp) if resp else default_choice
            if 1 <= choice <= len(files):
                return files[choice - 1]
        except Exception:
            pass
        print('Invalid selection. Try again.')


def prompt_interpolation():
    default = 'y'
    while True:
        resp = input(f'Can I interpolate between values as needed? (y/n) [{default}]: ').strip().lower()
        if not resp:
            resp = default
        if resp in ('y', 'n'):
            return resp == 'y'
        print('Please enter y or n.')


def read_spreadsheet(filepath):
    try:
        # Always read from Sheet1, header in row 3, data starts at row 4
        df = pd.read_excel(filepath, sheet_name='Sheet1', header=2)
        # Only keep rows from row 4 onward (header=2 means pandas row 0 is Excel row 3)
        df = df.dropna(how='all')
        # Map columns by Excel column letters
        # CK: Battery Display SOC [%], CE: Power [W(DC)], H: Time [hh:mm:ss.0], CF: Energy Charged [Wh(DC)]
        col_soc = 'Battery Display SOC [%]'
        col_power = 'Power [W(DC)]'
        col_time = 'Time [hh:mm:ss.0]'
        col_energy = 'Energy Charged [Wh(DC)]'
        for col in [col_soc, col_power, col_time, col_energy]:
            if col not in df.columns:
                print(f'Missing required column: {col}')
                print(f'Columns found: {list(df.columns)}')
                sys.exit(1)
        # print(f"Read {len(df)} rows from Sheet1 with columns: {list(df.columns)}")
        return df[[col_soc, col_power, col_time, col_energy]].copy()
    except Exception as e:
        print(f'Error reading spreadsheet: {e}')
        sys.exit(1)


def process_data(df, interpolate):
    # Use fixed column names as per user specification
    col_soc = 'Battery Display SOC [%]'
    col_power = 'Power [W(DC)]'
    col_time = 'Time [hh:mm:ss.0]'
    col_energy = 'Energy Charged [Wh(DC)]'
    soc = df[col_soc].astype(float)
    power = df[col_power].astype(float) / 1000.0  # Convert W to kW
    # Convert time from hh:mm:ss.0 to minutes
    def time_to_minutes(t):
        if pd.isnull(t):
            return np.nan
        parts = str(t).split(':')
        if len(parts) == 3:
            h, m, s = parts
            return float(h) * 60 + float(m) + float(s) / 60
        return np.nan
    time = df[col_time].apply(time_to_minutes)
    energy = df[col_energy].astype(float) / 1000.0  # Convert Wh to kWh
    # Scale SOC to 0-100 if needed
    if soc.max() <= 1.0:
        soc = soc * 100
    # Remove duplicate SOCs, keep last occurrence
    dedup = pd.DataFrame({'soc': soc, 'power': power, 'time': time, 'energy': energy})
    dedup = dedup.drop_duplicates(subset='soc', keep='last').sort_values('soc')
    soc = dedup['soc'].values
    power = dedup['power'].values
    time = dedup['time'].values
    energy = dedup['energy'].values

    # Extrapolation for missing SOCs at start/end
    soc_min, soc_max = soc[0], soc[-1]
    missing_low = int(soc_min) > 0
    missing_high = int(soc_max) < 100
    # Prompt user if extrapolation is needed
    if missing_low or missing_high:
        msg = []
        if missing_low:
            msg.append(f"{int(soc_min)}% and below")
        if missing_high:
            msg.append(f"{int(soc_max)}% and above")
        print(f"Data missing for SOC: {', '.join(msg)}.")
        default = 'y'
        resp = input(f"Should I extrapolate outside of the provided data to give a fuller picture? (y/n) [{default}]: ").strip().lower()
        if not resp:
            resp = default
        if resp != 'y':
            print('Aborting due to missing SOC range.')
            sys.exit(1)
    # Extrapolate lower end: energy first, then time, then power (power = dE/dt for each step)
    if missing_low:
        first_soc = soc[0]
        n_missing = int(first_soc)
        # Estimate total battery capacity
        usable_fraction = (100 - first_soc) / 100.0
        if usable_fraction > 0:
            total_capacity = energy[-1] / usable_fraction
        else:
            total_capacity = energy[-1]
        missing_energy = total_capacity - energy[-1]
        # Extrapolate energy: linearly from 0 to missing_energy
        extra_soc = np.arange(0, int(first_soc))
        extra_energy = np.linspace(0, missing_energy, n_missing, endpoint=False)
        # Calculate time for each step: use average power of first measured segment if possible, else use first measured power
        # But now, time is calculated so that power = dE/dt for each step
        # We'll use the average dE per step and the average power of the first measured segment
        # For more physical meaning, we can use the first measured dE and dt as a template
        if len(energy) > 1 and len(time) > 1:
            measured_dE = energy[1] - energy[0]
            measured_dt = time[1] - time[0]
            template_power = measured_dE / (measured_dt / 60) if measured_dt > 0 else power[0]
        else:
            template_power = power[0]
        # For each extrapolated step, calculate dE and assign a time so that power = dE/dt
        dE = np.diff(np.concatenate(([0], extra_energy)))  # kWh per step
        # Use template_power for each step to get dt, then cumulative sum for time
        extra_dt = np.where(template_power > 0, dE / template_power * 60, 0)  # minutes per step
        extra_time = np.cumsum(extra_dt)
        # Now, for each step, power = dE / (dt/60)
        with np.errstate(divide='ignore', invalid='ignore'):
            extra_power = np.where(extra_dt > 0, dE / (extra_dt / 60), 0)
            extra_power = np.nan_to_num(extra_power, nan=0.0, posinf=0.0, neginf=0.0)
        # Shift all existing energy up by missing_energy
        energy = energy + missing_energy
        # Shift all existing time up by total extrapolated time
        total_extrapolated_time = extra_time[-1] if len(extra_time) > 0 else 0
        time = time + total_extrapolated_time
        # Concatenate
        soc = np.concatenate([extra_soc, soc])
        power = np.concatenate([extra_power, power])
        energy = np.concatenate([extra_energy, energy])
        time = np.concatenate([extra_time, time])
    # Extrapolate upper end
    if missing_high:
        last_soc = soc[-1]
        last_power = power[-1]
        last_energy = energy[-1]
        last_time = time[-1]
        n_missing = 100 - int(last_soc)
        if n_missing > 0:
            # Estimate energy/time per %SOC for upper end
            if len(soc) > 1:
                delta_e = energy[-1] - energy[-2]
                delta_t = time[-1] - time[-2]
            else:
                delta_e = last_power / 60
                delta_t = 1
            extra_soc = np.arange(int(last_soc)+1, 101)
            extra_power = np.full_like(extra_soc, last_power, dtype=float)
            extra_energy = np.linspace(last_energy, last_energy + delta_e * n_missing, n_missing+1)[1:]
            extra_time = np.linspace(last_time, last_time + delta_t * n_missing, n_missing+1)[1:]
            soc = np.concatenate([soc, extra_soc])
            power = np.concatenate([power, extra_power])
            energy = np.concatenate([energy, extra_energy])
            time = np.concatenate([time, extra_time])
    # Interpolate/extrapolate to integer SOC 0-100
    soc_grid = np.arange(0, 101)
    power_interp = np.interp(soc_grid, soc, power)
    time_interp = np.interp(soc_grid, soc, time)
    if energy is not None:
        energy_interp = np.interp(soc_grid, soc, energy)
    else:
        # Estimate energy as cumulative sum of (power * delta_time)
        energy_interp = np.zeros_like(soc_grid, dtype=float)
        for i in range(1, len(soc_grid)):
            dt = time_interp[i] - time_interp[i-1]
            avg_power = (power_interp[i] + power_interp[i-1]) / 2
            energy_interp[i] = energy_interp[i-1] + avg_power * dt / 60  # kWh
    return soc_grid, power_interp, time_interp, energy_interp


def build_xml(soc, power, time, energy, vehicle_name, vehicle_range, range_speed):
    vehicle = Element('vehicle')
    metadata = SubElement(vehicle, 'metadata')
    SubElement(metadata, 'name').text = vehicle_name
    SubElement(metadata, 'range_miles').text = str(vehicle_range)
    SubElement(metadata, 'range_speed_mph').text = str(range_speed)
    curve = SubElement(vehicle, 'charging_curve', unit_time='minutes', unit_power='kW', unit_energy='kWh')
    for s, p, t, e in zip(soc, power, time, energy):
        SubElement(curve, 'point', soc=str(int(s)), power=f'{p:.0f}', time=f'{t:.2f}', energy=f'{e:.1f}')
    return vehicle


def main():
    files = list_excel_files()
    fname = prompt_file_selection(files)
    df = read_spreadsheet(os.path.join(INPUT_DIR, fname))
    # Get vehicle name from cell B3
    import openpyxl
    wb = openpyxl.load_workbook(os.path.join(INPUT_DIR, fname), data_only=True)
    ws = wb.active
    detected_name = ws['B3'].value if ws['B3'].value else ''
    if detected_name:
        print(f"Detected vehicle name from spreadsheet: {detected_name}")
        vehicle_name = input(f"Vehicle name [{detected_name}]: ").strip()
        if not vehicle_name:
            vehicle_name = detected_name
    else:
        default_vehicle = ''
        vehicle_name = input(f'Vehicle name [{default_vehicle}]: ').strip()
        if not vehicle_name:
            vehicle_name = default_vehicle
    vehicle_range = input('Range (miles): ')
    interpolate = prompt_interpolation()
    range_speed = 70
    soc, power, time, energy = process_data(df, interpolate)
    xml_root = build_xml(soc, power, time, energy, vehicle_name, vehicle_range, range_speed)
    # Sanitize vehicle name for filename
    import re
    safe_vehicle_name = vehicle_name.strip().replace('>', 'm').replace('<', 'l')
    safe_vehicle_name = re.sub(r'[^\w\-_\. ]', '_', safe_vehicle_name)
    outname = safe_vehicle_name + '.xml'
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
    outpath = os.path.join(OUTPUT_DIR, outname)

    # Pretty-print XML and use double quotes in declaration
    import xml.dom.minidom
    rough_string = ElementTree(xml_root).write(outpath, encoding='utf-8', xml_declaration=True)
    # Read back the file, pretty-print, and overwrite
    with open(outpath, 'rb') as f:
        xml_bytes = f.read()
    dom = xml.dom.minidom.parseString(xml_bytes)
    pretty_xml = dom.toprettyxml(indent='  ', encoding='utf-8')
    # Replace single quotes with double quotes in the XML declaration
    pretty_xml_str = pretty_xml.decode('utf-8')
    if pretty_xml_str.startswith("<?xml version='1.0' encoding='utf-8'?>"):
        pretty_xml_str = pretty_xml_str.replace("<?xml version='1.0' encoding='utf-8'?>", "<?xml version=\"1.0\" encoding=\"utf-8\"?>", 1)
    with open(outpath, 'w', encoding='utf-8') as f:
        f.write(pretty_xml_str)
    print(f'XML written to {outpath}')

if __name__ == '__main__':
    main()
